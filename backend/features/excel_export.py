from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io


def export_pl_excel(pl_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.append(["Profit & Loss Statement"])
    ws.append([f"Period: {pl_data.get('from_date')} to {pl_data.get('to_date')}"])
    ws.append([])
    rows = [
        ("Total Sales", pl_data.get("total_sales", 0)),
        ("Total Purchases", pl_data.get("total_purchases", 0)),
        ("Gross Profit", pl_data.get("gross_profit", 0)),
        ("Total Expenses", pl_data.get("total_expenses", 0)),
        ("Net Profit", pl_data.get("net_profit", 0)),
        ("Profit Margin (%)", pl_data.get("profit_margin", 0)),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws["A1"].font = Font(bold=True, size=14)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_gst_excel(gst_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Summary"
    ws.append(["GST Summary"])
    ws.append([f"Month: {gst_data.get('month')}/{gst_data.get('year')}"])
    ws.append([])
    rows = [
        ("Taxable Sales", gst_data.get("total_taxable_sales", 0)),
        ("CGST Collected", gst_data.get("cgst_collected", 0)),
        ("SGST Collected", gst_data.get("sgst_collected", 0)),
        ("IGST Collected", gst_data.get("igst_collected", 0)),
        ("Total GST Collected", gst_data.get("total_gst_collected", 0)),
        ("GST Paid (Purchases)", gst_data.get("total_gst_paid_on_purchases", 0)),
        ("Net GST Liability", gst_data.get("net_gst_liability", 0)),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws["A1"].font = Font(bold=True, size=14)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
